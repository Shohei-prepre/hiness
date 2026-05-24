"""
Excelエクスポートスクリプト
enriched_companies.csv または safe.csv を読み込み、
指定フォーマットの Excel ファイルに変換する

実行: python export_excel.py
    : python export_excel.py --input data/safe.csv --output 営業リスト.xlsx
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from config import ENRICHED_CSV, SAFE_CSV, DATA_DIR

# ────────────────────────────────
# スタイル定義（スクリーンショットに合わせた配色）
# ────────────────────────────────
HEADER_BG    = "1F3864"   # 濃紺
HEADER_FONT  = "FFFFFF"   # 白
NO_BG        = "1F3864"   # No.列も濃紺
ROW_ODD_BG   = "FFFFFF"   # 奇数行：白
ROW_EVEN_BG  = "EEF2F7"  # 偶数行：薄青グレー
BORDER_COLOR = "B8C4D6"

# 列定義：(ヘッダー表示名, CSVカラム名 or None, 列幅)
COLUMNS = [
    ("No.",                   None,           5),
    ("企業名",                "company_name", 22),
    ("サービス名",            "service_name", 22),
    ("カテゴリ",              "category",     18),
    ("本社所在地",            "location",     18),
    ("公式サイト（問い合わせ窓口）", "form_url", 40),
    ("備考・特徴",            "notes",        30),
]


def make_border() -> Border:
    """罫線スタイルを生成する"""
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_header(ws, col_idx: int, text: str) -> None:
    """ヘッダーセルにスタイルを適用する"""
    cell = ws.cell(row=1, column=col_idx, value=text)
    cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    cell.font      = Font(bold=True, color=HEADER_FONT, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = make_border()


def apply_data_cell(ws, row: int, col: int, value, is_even: bool, is_no: bool = False) -> None:
    """データセルにスタイルを適用する"""
    cell = ws.cell(row=row, column=col, value=value)

    bg = NO_BG if is_no else (ROW_EVEN_BG if is_even else ROW_ODD_BG)
    cell.fill = PatternFill("solid", fgColor=bg)

    font_color = HEADER_FONT if is_no else "000000"
    cell.font  = Font(size=9, color=font_color, bold=is_no)

    halign = "center" if is_no else "left"
    cell.alignment = Alignment(
        horizontal=halign, vertical="center", wrap_text=True
    )
    cell.border = make_border()


def load_csv(path: str) -> pd.DataFrame:
    """CSVを読み込み、不足カラムを空欄で補完して返す"""
    df = pd.read_csv(path, encoding="utf-8-sig", dtype=str).fillna("")

    # カラム名のマッピング（エージェントごとに列名が違う場合を吸収）
    rename_map = {
        "official_url": "form_url",   # agent1のみの場合はofficial_urlをform_urlに充てる
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns and v not in df.columns})

    # 不足カラムを空欄で追加
    for _, csv_col, _ in COLUMNS:
        if csv_col and csv_col not in df.columns:
            df[csv_col] = ""

    # カテゴリを固定値で埋める
    if "category" in df.columns:
        df["category"] = df["category"].replace("", "求人広告代理店")
    else:
        df["category"] = "求人広告代理店"

    # sourceがあればnotesに充てる
    if "notes" not in df.columns or df["notes"].eq("").all():
        if "source" in df.columns:
            df["notes"] = df["source"]

    return df


def export(input_path: str, output_path: str) -> None:
    """CSVを読み込んでExcelに書き出す"""
    if not os.path.exists(input_path):
        print(f"❌ 入力ファイルが見つかりません: {input_path}")
        return

    df = load_csv(input_path)
    print(f"[読込] {input_path} ({len(df)} 件)")

    wb = Workbook()
    ws = wb.active
    ws.title = "営業リスト"

    # ヘッダー行を書き込む
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        apply_header(ws, col_idx, header)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ヘッダー行の高さ
    ws.row_dimensions[1].height = 30

    # データ行を書き込む
    for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
        excel_row = row_idx + 1     # 1行目はヘッダー
        is_even   = row_idx % 2 == 0
        ws.row_dimensions[excel_row].height = 22

        for col_idx, (_, csv_col, _) in enumerate(COLUMNS, start=1):
            is_no = col_idx == 1
            if is_no:
                value = row_idx
            elif csv_col:
                value = row.get(csv_col, "")
            else:
                value = ""

            apply_data_cell(ws, excel_row, col_idx, value, is_even, is_no)

    # ウィンドウ枠の固定（ヘッダー行と No. 列）
    ws.freeze_panes = "B2"

    # オートフィルター
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"[完了] Excelエクスポート: {output_path} ({len(df)} 社)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="CSVを営業リストExcelに変換する")
    parser.add_argument(
        "--input",
        default=ENRICHED_CSV,
        help=f"入力CSVパス（デフォルト: {ENRICHED_CSV}）",
    )
    parser.add_argument(
        "--output",
        default=os.path.join(DATA_DIR, "ハイネス_営業リスト.xlsx"),
        help="出力Excelパス",
    )
    args = parser.parse_args()
    export(args.input, args.output)
